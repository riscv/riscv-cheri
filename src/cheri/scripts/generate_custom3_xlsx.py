#!/usr/bin/env python3
## WARNING: This script is almost entirely AI generated. The output looks correct, but this code might not be
import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from rvy_instruction_encodings import (
    get_custom3_insts,
    Custom3Funct3,
    MajorOpcode,
    IType,
    AMOType,
    InsnBitsCell,
    CellType,
    TableEntry,
)

THICK_SIDE = Side(style="thick", color="000000")

# Style constants
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
COLOR_HEADER = "DDEBF7"
COLOR_RED = PatternFill(start_color="FCE4D6", fill_type="solid")


class SeparatorRow(TableEntry):
    def __init__(self, boxes: list[InsnBitsCell], name: str = "", is_header: bool = False, comment=""):
        super().__init__(name, ext=None, comment=comment)
        self.boxes = boxes
        self.is_header = is_header


def render_row(entry, ws, row_idx, default_op, merge_box):
    from rvy_instruction_encodings import Instruction
    if isinstance(entry, SeparatorRow):
        for cell in entry.boxes:
            fill_color = cell.cell_type.fill_color
            if entry.is_header:
                fill_color = COLOR_HEADER
            merge_box(row_idx, cell.start, cell.end, cell.value_for_encoding_overview, fill_color=fill_color)
        if entry.is_header:
            # Apply header font and alignment to the whole row
            for col in range(1, 36):
                ws.cell(row_idx, col).font = HEADER_FONT
                ws.cell(row_idx, col).alignment = CENTER_ALIGN
    elif isinstance(entry, Instruction):
        for cell in entry.cells:
            merge_box(row_idx, cell.start, cell.end, cell.value_for_encoding_overview, is_fixed=cell.is_fixed, cell_type=cell.cell_type)


def generate_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custom-3 Proposal"

    def merge_box(row, start, end, text, is_fixed=True, fill_color=None, cell_type=None):
        # Rename {cd}/{cs1}/{cs2} to rd (YLEN)/rs1 (YLEN)/rs2 (YLEN) for XLSX
        text = str(text).replace("{cd}", "rd (YLEN)").replace("{cs1}", "rs1 (YLEN)").replace("{cs2}", "rs2 (YLEN)")
        s_col, e_col = 2 + (31 - start), 2 + (31 - end)
        ws.cell(row, s_col).value = text
        for c in range(s_col, e_col + 1):
            color_str = fill_color if fill_color else (cell_type.fill_color if cell_type else "FFFFFF")
            ws.cell(row, c).fill = PatternFill(start_color=color_str, fill_type="solid")
            ws.cell(row, c).alignment = CENTER_ALIGN
            ws.cell(row, c).border = Border(
                left=(THICK_SIDE if c == s_col else None),
                right=(THICK_SIDE if c == e_col else None),
                top=THICK_SIDE,
                bottom=THICK_SIDE,
            )
        if s_col < e_col:
            ws.merge_cells(start_row=row, start_column=s_col, end_row=row, end_column=e_col)

    ws.append(["Instruction Name"] + [str(i) for i in range(31, -1, -1)] + ["Extension", "Comments"])
    for col in range(1, 36):
        ws.cell(1, col).font, ws.cell(1, col).alignment = HEADER_FONT, CENTER_ALIGN
        ws.cell(1, col).border = Border(
            left=(THICK_SIDE if col in (1, 34, 35) else None),
            right=(THICK_SIDE if col in (1, 33, 34, 35) else None),
            top=THICK_SIDE,
            bottom=THICK_SIDE,
        )

    rvy_insts = get_custom3_insts()
    insts = (
        [
            SeparatorRow(
                name="Regular 3OP",
                is_header=True,
                boxes=[
                    InsnBitsCell("FUNCT7", 31, 25, cell_type=CellType.HEADER),
                    InsnBitsCell("RS2", 24, 20, cell_type=CellType.HEADER),
                    InsnBitsCell("RS1", 19, 15, cell_type=CellType.HEADER),
                    InsnBitsCell("FUNCT3", 14, 12, cell_type=CellType.HEADER),
                    InsnBitsCell("RD", 11, 7, cell_type=CellType.HEADER),
                    InsnBitsCell("MAJOR OPCODE", 6, 0, cell_type=CellType.HEADER),
                ],
                comment="bit[0]: rd is YLEN, bit[1]: rs1 is YLEN, bit[2]: rs2 is YLEN\nto see register access width during decode",
            )
        ]
        + rvy_insts.regular_3op_insns
        + [
            SeparatorRow(
                name="Regular 2OP",
                is_header=True,
                boxes=[
                    InsnBitsCell("FUNCT7", 31, 25, cell_type=CellType.HEADER),
                    InsnBitsCell("FUNCT5", 24, 20, cell_type=CellType.HEADER),
                    InsnBitsCell("RS1", 19, 15, cell_type=CellType.HEADER),
                    InsnBitsCell("FUNCT3", 14, 12, cell_type=CellType.HEADER),
                    InsnBitsCell("RD", 11, 7, cell_type=CellType.HEADER),
                    InsnBitsCell("MAJOR OPCODE", 6, 0, cell_type=CellType.HEADER),
                ],
                comment="funct7 bit[0]: rd is YLEN, bit[1]: rs1 is YLEN, bit[2]=0, following 3OP approach",
            )
        ]
        + rvy_insts.regular_2op_insns
        + [
            SeparatorRow(
                name="I-Type",
                is_header=True,
                boxes=[
                    InsnBitsCell("IMM", 31, 25, cell_type=CellType.HEADER),
                    InsnBitsCell("RS2/IMM", 24, 20, cell_type=CellType.HEADER),
                    InsnBitsCell("RS1", 19, 15, cell_type=CellType.HEADER),
                    InsnBitsCell("FUNCT3", 14, 12, cell_type=CellType.HEADER),
                    InsnBitsCell("RD/IMM", 11, 7, cell_type=CellType.HEADER),
                    InsnBitsCell("MAJOR OPCODE", 6, 0, cell_type=CellType.HEADER),
                ],
            ),
            rvy_insts.yaddi,
            rvy_insts.ly,
            rvy_insts.sy,
            SeparatorRow(
                name="Atomic Memory Operations",
                is_header=True,
                boxes=[
                    InsnBitsCell("FUNCT5", 31, 27, cell_type=CellType.HEADER),
                    InsnBitsCell("AQ", 26, 26, cell_type=CellType.HEADER),
                    InsnBitsCell("RL", 25, 25, cell_type=CellType.HEADER),
                    InsnBitsCell("RS2", 24, 20, cell_type=CellType.HEADER),
                    InsnBitsCell("RS1", 19, 15, cell_type=CellType.HEADER),
                    InsnBitsCell("FUNCT3", 14, 12, cell_type=CellType.HEADER),
                    InsnBitsCell("RD", 11, 7, cell_type=CellType.HEADER),
                    InsnBitsCell("MAJOR OPCODE", 6, 0, cell_type=CellType.HEADER),
                ],
            ),
        ]
        + rvy_insts.amo_insns
        + [
            SeparatorRow(
                name="Misc. Imm/Shift formats",
                is_header=True,
                boxes=[
                    InsnBitsCell("FUNCT/IMM/RS2", 31, 20, cell_type=CellType.HEADER),
                    InsnBitsCell("RS1", 19, 15, cell_type=CellType.HEADER),
                    InsnBitsCell("FUNCT3", 14, 12, cell_type=CellType.HEADER),
                    InsnBitsCell("RD/IMM", 11, 7, cell_type=CellType.HEADER),
                    InsnBitsCell("MAJOR OPCODE", 6, 0, cell_type=CellType.HEADER),
                ],
            )
        ]
        + rvy_insts.misc_insns
    )

    for inst in insts:
        row_idx = ws.max_row + 1
        ws.append([inst.name] + [""] * 32 + [inst.ext, inst.comment])
        ws.cell(row_idx, 34).border = Border(left=THICK_SIDE, right=THICK_SIDE, top=THICK_SIDE, bottom=THICK_SIDE)
        ws.cell(row_idx, 35).border = Border(left=THICK_SIDE, right=THICK_SIDE, top=THICK_SIDE, bottom=THICK_SIDE)
        ws.cell(row_idx, 1).border = Border(left=THICK_SIDE, right=THICK_SIDE, top=THICK_SIDE, bottom=THICK_SIDE)

        render_row(inst, ws, row_idx, MajorOpcode.RVY_A, merge_box)

        if "RESERVED" in inst.name or "UNALLOCATED" in inst.name:
            for col in range(1, 34):
                cell = ws.cell(row_idx, col)
                # COLOR_VAR is FFF2CC. Unset or yellow fields should become red.
                if cell.fill.start_color.rgb in ("00000000", "FFFFF2CC"):
                    cell.fill = COLOR_RED

    ws.column_dimensions["A"].width = 25
    for i in range(2, 34):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 4
    ws.column_dimensions[openpyxl.utils.get_column_letter(34)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(35)].width = 30

    # --- Add Allocations Tables sheet ---
    ws_tables = wb.create_sheet("Allocations Tables")
    thin_side = Side(style="thin")

    def write_table_header(title, headers):
        ws_tables.append([])
        row_idx = ws_tables.max_row + 1
        ws_tables.cell(row_idx, 1, title).font = Font(bold=True, size=14)
        ws_tables.append(headers)
        header_row = ws_tables.max_row
        for col in range(1, len(headers) + 1):
            cell = ws_tables.cell(header_row, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def append_data_row(row_data):
        ws_tables.append(row_data)
        row_idx = ws_tables.max_row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_tables.cell(row_idx, col_idx)
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            if col_idx == 1:
                cell.font = Font(bold=True)

    # 1. Funct3 Allocations
    f3_map = Custom3Funct3.get_map()
    write_table_header("1. Funct3 Allocations", ["funct3", "Instruction(s)"])
    for i in range(8):
        append_data_row([format(i, "03b"), f3_map[i]])

    # 2. R-Type 3-Operand
    r3_grouped = {}
    for i in rvy_insts.regular_3op_insns:
        if isinstance(i.f7.val, int):
            r3_grouped.setdefault(i.f7.val, []).append(i.name)

    headers = ["funct7[6:3] \\ funct7[2:0]"] + [format(c, "03b") for c in range(8)]
    write_table_header("2. R-Type 3-Operand (funct3=000)", headers)
    for r in range(16):
        row_data = [f"{format(r, '04b')} (0x{r:X})"]
        for c in range(8):
            f7 = (r << 3) | c
            if f7 == 0x7F:
                name = "1OP/2OP"
            else:
                names = r3_grouped.get(f7, [])
                name = "/".join(names)
            row_data.append(name)
        append_data_row(row_data)

    # 3. R-Type 1/2-Operand
    r2_by_f7 = {}
    for i in rvy_insts.regular_2op_insns:
        if isinstance(i.rs2.val, int) and hasattr(i, "f7") and isinstance(i.f7.val, int):
            r2_by_f7.setdefault(i.f7.val, []).append(i)

    table_num = 3
    for f7_val, r2_insts in sorted(r2_by_f7.items()):
        headers = ["rs2[2:0] \\ rs2[4:3]", "00", "01", "10", "11"]
        write_table_header(f"{table_num}. R-Type 1/2-Operand (funct3=000, funct7={format(f7_val, '07b')})", headers)

        r2_grouped = {}
        for i in r2_insts:
            r2_grouped.setdefault(i.rs2.val, []).append(i.name)

        for r in range(8):
            row_data = [format(r, "03b")]
            for c in range(4):
                rs2 = (c << 3) | r
                names = r2_grouped.get(rs2, [])
                name = "/".join(names)
                row_data.append(name)
            append_data_row(row_data)
        table_num += 1

    # 4. AMO Sub-opcode Allocations
    amo_grouped = {}
    for i in rvy_insts.amo_insns:
        matching_f7s = []
        if isinstance(i, AMOType) and isinstance(i.f5_fixed.val, int):
            base_f7 = i.f5_fixed.val << 2
            for offset in range(4):
                matching_f7s.append(base_f7 + offset)
        for f7 in matching_f7s:
            amo_grouped.setdefault(f7, []).append(i.name)

    headers = ["funct7[6:3] \\ funct7[2:0]"] + [format(c, "03b") for c in range(8)]
    write_table_header(f"{table_num}. AMO Sub-opcode Allocations (funct3=100)", headers)
    for r in range(16):
        row_data = [f"{format(r, '04b')} (0x{r:X})"]
        for c in range(8):
            f7 = (r << 3) | c
            names = amo_grouped.get(f7, [])
            name = "/".join(names)
            row_data.append(name)
        append_data_row(row_data)
    table_num += 1

    # 5. MISC Sub-opcode Allocations
    misc_grouped = {}
    misc_f3 = Custom3Funct3.MISC
    for i in rvy_insts.misc_insns:
        if i.op.val == MajorOpcode.RVY_A and i.f3.val == misc_f3:
            matching_f7s = []
            if isinstance(i, IType) and hasattr(i, "imm_high"):
                fixed_val = i.imm_high.val
                L = len(fixed_val)
                top_val = int(fixed_val, 2)
                if L <= 7:
                    shift = 7 - L
                    base_f7 = top_val << shift
                    for offset in range(1 << shift):
                        f7 = base_f7 + offset
                        if "shamt=XLEN" in i.name and f7 != 0:
                            continue
                        matching_f7s.append(f7)

            for f7 in matching_f7s:
                misc_grouped.setdefault(f7, []).append(i.name)

    headers = ["funct7[6:3] \\ funct7[2:0]"] + [format(c, "03b") for c in range(8)]
    write_table_header(f"{table_num}. MISC Sub-opcode Allocations (funct3=101)", headers)
    for r in range(16):
        row_data = [f"{format(r, '04b')} (0x{r:X})"]
        for c in range(8):
            f7 = (r << 3) | c
            names = misc_grouped.get(f7, [])
            name = "/".join(names)
            row_data.append(name)
        append_data_row(row_data)

    def print_sheet_to_stdout(sheet, title):
        merged_cells_info = {}
        skip_cells = set()
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left = (min_row, min_col)
            merged_cells_info[top_left] = (max_row - min_row + 1, max_col - min_col + 1)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != top_left:
                        skip_cells.add((r, c))

        max_col_idx = sheet.max_column
        max_row_idx = sheet.max_row

        col_widths = {}
        for col in range(1, max_col_idx + 1):
            max_length = 0
            for row in range(1, max_row_idx + 1):
                if (row, col) not in skip_cells and (row, col) not in merged_cells_info:
                    val = sheet.cell(row, col).value
                    if val is not None:
                        for line in str(val).split("\n"):
                            max_length = max(max_length, len(line))
            col_widths[col] = max_length + 2

        for top_left, (r_span, c_span) in merged_cells_info.items():
            r, c = top_left
            val = sheet.cell(r, c).value
            if val is not None:
                max_len = max((len(line) for line in str(val).split("\n")), default=0)
                current_width = sum(col_widths.get(c + i, 2) for i in range(c_span)) + 3 * (c_span - 1)
                if max_len + 2 > current_width:
                    col_widths[c + c_span - 1] = col_widths.get(c + c_span - 1, 2) + (max_len + 2 - current_width)

        col_widths_list = [col_widths.get(i, 2) for i in range(1, max_col_idx + 1)]
        print_cols = min(max_col_idx, 35)
        total_width = sum(col_widths_list[:print_cols]) + 3 * print_cols

        print("\n" + "=" * total_width)
        print(title.center(total_width))
        print("=" * total_width)

        for row in range(1, max_row_idx + 1):
            row_cells = []
            c = 1
            is_empty_row = True
            while c <= print_cols:
                if (row, c) in skip_cells:
                    row_cells.append(("", c, 1))
                    c += 1
                    continue

                val = sheet.cell(row, c).value
                if val is not None:
                    is_empty_row = False
                text = str(val if val is not None else "")

                span = 1
                if (row, c) in merged_cells_info:
                    _, c_span = merged_cells_info[(row, c)]
                    span = min(c_span, print_cols - c + 1)

                row_cells.append((text, c, span))
                c += span

            if is_empty_row:
                print("-" * total_width)
                continue

            split_row = [text.split("\n") for text, _, _ in row_cells]
            max_lines = max((len(lines) for lines in split_row), default=1)

            for line_idx in range(max_lines):
                row_str = ""
                for i, (text, col_start, span) in enumerate(row_cells):
                    w = sum(col_widths_list[col_start - 1 : col_start - 1 + span]) + 3 * (span - 1)
                    val_str = split_row[i][line_idx] if line_idx < len(split_row[i]) else ""
                    if span > 1:
                        row_str += val_str.center(w) + " | "
                    else:
                        row_str += val_str.ljust(w) + " | "
                print(row_str)
        print("=" * total_width + "\n")

    # Auto-adjust column widths for ws_tables
    for column_cells in ws_tables.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_tables.column_dimensions[column_letter].width = max_length + 2

    print_sheet_to_stdout(ws, "Custom-3 Bit Allocations")
    print_sheet_to_stdout(ws_tables, "Allocations Tables Overview")

    wb.save("Custom3_RVY_Proposal.xlsx")
    print("Successfully generated Custom3_RVY_Proposal.xlsx")


if __name__ == "__main__":
    generate_xlsx()
